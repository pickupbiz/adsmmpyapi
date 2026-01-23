"""
Report generation utilities for PDF and Excel exports.

Provides:
- PDF report generation using ReportLab
- Excel report generation using openpyxl
- CSV export functionality
"""
import io
import os
from datetime import datetime, date, timedelta
from typing import List, Dict, Any, Optional, Union
from decimal import Decimal
from uuid import uuid4

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, Image, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

from app.core.config import settings


class ReportGenerator:
    """Base class for report generation."""
    
    def __init__(self):
        self.reports_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_filename(self, report_name: str, extension: str) -> str:
        """Generate unique filename for report."""
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        unique_id = str(uuid4())[:8]
        return f"{report_name}_{timestamp}_{unique_id}.{extension}"
    
    def get_report_path(self, filename: str) -> str:
        """Get full path for report file."""
        return os.path.join(self.reports_dir, filename)


class PDFReportGenerator(ReportGenerator):
    """PDF report generation using ReportLab."""
    
    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            raise RuntimeError(
                "ReportLab is not installed. Install with `pip install reportlab` to enable PDF reports."
            )
        super().__init__()
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10
        ))
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        ))
    
    def create_header(self, title: str, subtitle: str = None) -> List:
        """Create report header elements."""
        elements = []
        elements.append(Paragraph(title, self.styles['ReportTitle']))
        if subtitle:
            elements.append(Paragraph(subtitle, self.styles['Normal']))
        elements.append(Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
            self.styles['Normal']
        ))
        elements.append(Spacer(1, 20))
        return elements
    
    def create_table(
        self,
        headers: List[str],
        data: List[List[Any]],
        col_widths: List[float] = None
    ) -> Table:
        """Create a styled table."""
        table_data = [headers] + data
        
        if col_widths:
            table = Table(table_data, colWidths=col_widths)
        else:
            table = Table(table_data)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ])
        table.setStyle(style)
        return table
    
    def generate_po_report(
        self,
        pos: List[Dict[str, Any]],
        title: str = "Purchase Order Report"
    ) -> str:
        """Generate PO report PDF."""
        filename = self.generate_filename("po_report", "pdf")
        filepath = self.get_report_path(filename)
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = self.create_header(title, "Aerospace Materials Management System")
        
        # Summary section
        elements.append(Paragraph("Summary", self.styles['SectionTitle']))
        total_pos = len(pos)
        total_value = sum(float(po.get('total_amount', 0)) for po in pos)
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Purchase Orders", str(total_pos)],
            ["Total Value", f"${total_value:,.2f}"],
        ]
        elements.append(self.create_table(
            summary_data[0],
            summary_data[1:],
            col_widths=[3*inch, 2*inch]
        ))
        elements.append(Spacer(1, 20))
        
        # PO Details
        elements.append(Paragraph("Purchase Order Details", self.styles['SectionTitle']))
        
        headers = ["PO Number", "Supplier", "Date", "Status", "Amount"]
        data = []
        for po in pos:
            data.append([
                po.get('po_number', ''),
                po.get('supplier_name', '')[:20],
                str(po.get('order_date', ''))[:10],
                po.get('status', ''),
                f"${float(po.get('total_amount', 0)):,.2f}"
            ])
        
        if data:
            elements.append(self.create_table(
                headers,
                data,
                col_widths=[1.2*inch, 1.5*inch, 1*inch, 1.2*inch, 1.1*inch]
            ))
        
        doc.build(elements)
        return filepath
    
    def generate_material_report(
        self,
        materials: List[Dict[str, Any]],
        title: str = "Material Status Report"
    ) -> str:
        """Generate material status report PDF."""
        filename = self.generate_filename("material_report", "pdf")
        filepath = self.get_report_path(filename)
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = self.create_header(title, "Material Tracking Report")
        
        # Material Details
        headers = ["Material", "PO Ref", "Status", "Quantity", "Location"]
        data = []
        for mat in materials:
            data.append([
                mat.get('material_name', '')[:20],
                mat.get('po_number', 'N/A'),
                mat.get('status', ''),
                f"{float(mat.get('quantity', 0)):,.2f} {mat.get('unit', '')}",
                mat.get('location', 'N/A')
            ])
        
        if data:
            elements.append(self.create_table(
                headers,
                data,
                col_widths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.1*inch]
            ))
        
        doc.build(elements)
        return filepath
    
    def generate_supplier_performance_report(
        self,
        suppliers: List[Dict[str, Any]],
        title: str = "Supplier Performance Report"
    ) -> str:
        """Generate supplier performance report PDF."""
        filename = self.generate_filename("supplier_performance", "pdf")
        filepath = self.get_report_path(filename)
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = self.create_header(title, "Supplier Analytics")
        
        headers = ["Supplier", "Total POs", "On-Time %", "Quality %", "Score"]
        data = []
        for sup in suppliers:
            data.append([
                sup.get('supplier_name', '')[:25],
                str(sup.get('total_pos', 0)),
                f"{float(sup.get('on_time_delivery_rate', 0)):.1f}%",
                f"{float(sup.get('quality_acceptance_rate', 0)):.1f}%",
                f"{float(sup.get('performance_score', 0)):.1f}"
            ])
        
        if data:
            elements.append(self.create_table(
                headers,
                data,
                col_widths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch]
            ))
        
        doc.build(elements)
        return filepath
    
    def generate_inventory_report(
        self,
        inventory: List[Dict[str, Any]],
        title: str = "Inventory Status Report"
    ) -> str:
        """Generate inventory status report PDF."""
        filename = self.generate_filename("inventory_report", "pdf")
        filepath = self.get_report_path(filename)
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = self.create_header(title, "Current Stock Levels")
        
        headers = ["Material", "Current Qty", "Min Stock", "Reorder Lvl", "Status"]
        data = []
        for item in inventory:
            qty = float(item.get('quantity', 0))
            min_stock = float(item.get('minimum_stock', 0))
            status = "OK" if qty > min_stock else ("LOW" if qty > 0 else "OUT")
            
            data.append([
                item.get('material_name', '')[:20],
                f"{qty:,.2f}",
                f"{min_stock:,.2f}",
                f"{float(item.get('reorder_level', 0)):,.2f}",
                status
            ])
        
        if data:
            elements.append(self.create_table(
                headers,
                data,
                col_widths=[2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch]
            ))
        
        doc.build(elements)
        return filepath


class ExcelReportGenerator(ReportGenerator):
    """Excel report generation using openpyxl."""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl is not installed. Install with `pip install openpyxl` to enable Excel reports."
            )
        super().__init__()
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int):
        """Apply header style to a row."""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_po_report(
        self,
        pos: List[Dict[str, Any]],
        title: str = "Purchase Order Report"
    ) -> str:
        """Generate PO report Excel."""
        filename = self.generate_filename("po_report", "xlsx")
        filepath = self.get_report_path(filename)
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = title
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        ws_summary['A4'] = "Total POs"
        ws_summary['B4'] = len(pos)
        ws_summary['A5'] = "Total Value"
        ws_summary['B5'] = sum(float(po.get('total_amount', 0)) for po in pos)
        ws_summary['B5'].number_format = '$#,##0.00'
        
        # PO Details sheet
        ws_details = wb.create_sheet("PO Details")
        
        headers = ["PO Number", "Supplier", "Order Date", "Expected Delivery", 
                   "Status", "Priority", "Total Amount", "Created By"]
        
        for col, header in enumerate(headers, 1):
            ws_details.cell(row=1, column=col, value=header)
        self._apply_header_style(ws_details, 1, 1, len(headers))
        
        for row, po in enumerate(pos, 2):
            ws_details.cell(row=row, column=1, value=po.get('po_number', ''))
            ws_details.cell(row=row, column=2, value=po.get('supplier_name', ''))
            ws_details.cell(row=row, column=3, value=str(po.get('order_date', ''))[:10])
            ws_details.cell(row=row, column=4, value=str(po.get('expected_delivery_date', ''))[:10])
            ws_details.cell(row=row, column=5, value=po.get('status', ''))
            ws_details.cell(row=row, column=6, value=po.get('priority', ''))
            cell = ws_details.cell(row=row, column=7, value=float(po.get('total_amount', 0)))
            cell.number_format = '$#,##0.00'
            ws_details.cell(row=row, column=8, value=po.get('created_by', ''))
        
        self._auto_adjust_columns(ws_details)
        
        # Line Items sheet
        ws_lines = wb.create_sheet("Line Items")
        line_headers = ["PO Number", "Material", "Quantity", "Unit", "Unit Price", 
                        "Total Price", "Received Qty", "Status"]
        
        for col, header in enumerate(line_headers, 1):
            ws_lines.cell(row=1, column=col, value=header)
        self._apply_header_style(ws_lines, 1, 1, len(line_headers))
        
        row = 2
        for po in pos:
            for item in po.get('line_items', []):
                ws_lines.cell(row=row, column=1, value=po.get('po_number', ''))
                ws_lines.cell(row=row, column=2, value=item.get('material_name', ''))
                ws_lines.cell(row=row, column=3, value=float(item.get('quantity', 0)))
                ws_lines.cell(row=row, column=4, value=item.get('unit', ''))
                cell = ws_lines.cell(row=row, column=5, value=float(item.get('unit_price', 0)))
                cell.number_format = '$#,##0.00'
                cell = ws_lines.cell(row=row, column=6, value=float(item.get('total_price', 0)))
                cell.number_format = '$#,##0.00'
                ws_lines.cell(row=row, column=7, value=float(item.get('received_quantity', 0)))
                ws_lines.cell(row=row, column=8, value=item.get('status', ''))
                row += 1
        
        self._auto_adjust_columns(ws_lines)
        
        wb.save(filepath)
        return filepath
    
    def generate_material_report(
        self,
        materials: List[Dict[str, Any]],
        title: str = "Material Status Report"
    ) -> str:
        """Generate material status report Excel."""
        filename = self.generate_filename("material_report", "xlsx")
        filepath = self.get_report_path(filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Material Status"
        
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        headers = ["Material ID", "Material Name", "Barcode", "PO Number", 
                   "Status", "Quantity", "Unit", "Location", "Last Updated"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4, 1, len(headers))
        
        for row, mat in enumerate(materials, 5):
            ws.cell(row=row, column=1, value=mat.get('material_id', ''))
            ws.cell(row=row, column=2, value=mat.get('material_name', ''))
            ws.cell(row=row, column=3, value=mat.get('barcode', ''))
            ws.cell(row=row, column=4, value=mat.get('po_number', 'N/A'))
            ws.cell(row=row, column=5, value=mat.get('status', ''))
            ws.cell(row=row, column=6, value=float(mat.get('quantity', 0)))
            ws.cell(row=row, column=7, value=mat.get('unit', ''))
            ws.cell(row=row, column=8, value=mat.get('location', ''))
            ws.cell(row=row, column=9, value=str(mat.get('updated_at', ''))[:19])
        
        self._auto_adjust_columns(ws)
        wb.save(filepath)
        return filepath
    
    def generate_supplier_performance_report(
        self,
        suppliers: List[Dict[str, Any]],
        title: str = "Supplier Performance Report"
    ) -> str:
        """Generate supplier performance report Excel."""
        filename = self.generate_filename("supplier_performance", "xlsx")
        filepath = self.get_report_path(filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Performance Metrics"
        
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        headers = ["Supplier", "Code", "Total POs", "Completed", "Total Value",
                   "On-Time %", "Quality %", "Accuracy %", "Score", "Trend"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4, 1, len(headers))
        
        for row, sup in enumerate(suppliers, 5):
            ws.cell(row=row, column=1, value=sup.get('supplier_name', ''))
            ws.cell(row=row, column=2, value=sup.get('supplier_code', ''))
            ws.cell(row=row, column=3, value=sup.get('total_pos', 0))
            ws.cell(row=row, column=4, value=sup.get('completed_pos', 0))
            cell = ws.cell(row=row, column=5, value=float(sup.get('total_value', 0)))
            cell.number_format = '$#,##0.00'
            ws.cell(row=row, column=6, value=f"{float(sup.get('on_time_delivery_rate', 0)):.1f}%")
            ws.cell(row=row, column=7, value=f"{float(sup.get('quality_acceptance_rate', 0)):.1f}%")
            ws.cell(row=row, column=8, value=f"{float(sup.get('quantity_accuracy_rate', 0)):.1f}%")
            ws.cell(row=row, column=9, value=f"{float(sup.get('performance_score', 0)):.1f}")
            ws.cell(row=row, column=10, value=sup.get('performance_trend', 'stable'))
        
        self._auto_adjust_columns(ws)
        
        # Add chart if there's data
        if suppliers:
            ws_chart = wb.create_sheet("Performance Chart")
            
            # Copy data for chart
            ws_chart['A1'] = "Supplier"
            ws_chart['B1'] = "Score"
            for i, sup in enumerate(suppliers[:10], 2):  # Top 10
                ws_chart.cell(row=i, column=1, value=sup.get('supplier_name', '')[:15])
                ws_chart.cell(row=i, column=2, value=float(sup.get('performance_score', 0)))
            
            chart = BarChart()
            chart.title = "Supplier Performance Scores"
            chart.y_axis.title = "Score"
            
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=min(len(suppliers) + 1, 11))
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=min(len(suppliers) + 1, 11))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            
            ws_chart.add_chart(chart, "D1")
        
        wb.save(filepath)
        return filepath
    
    def generate_inventory_report(
        self,
        inventory: List[Dict[str, Any]],
        title: str = "Inventory Status Report"
    ) -> str:
        """Generate inventory report Excel."""
        filename = self.generate_filename("inventory_report", "xlsx")
        filepath = self.get_report_path(filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        headers = ["Material ID", "Material Name", "Current Qty", "Unit",
                   "Min Stock", "Reorder Level", "Location", "Status", "Pending PO"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=4, column=col, value=header)
        self._apply_header_style(ws, 4, 1, len(headers))
        
        for row, item in enumerate(inventory, 5):
            qty = float(item.get('quantity', 0))
            min_stock = float(item.get('minimum_stock', 0))
            status = "OK" if qty > min_stock else ("LOW" if qty > 0 else "OUT OF STOCK")
            
            ws.cell(row=row, column=1, value=item.get('material_id', ''))
            ws.cell(row=row, column=2, value=item.get('material_name', ''))
            ws.cell(row=row, column=3, value=qty)
            ws.cell(row=row, column=4, value=item.get('unit', ''))
            ws.cell(row=row, column=5, value=min_stock)
            ws.cell(row=row, column=6, value=float(item.get('reorder_level', 0)))
            ws.cell(row=row, column=7, value=item.get('location', ''))
            
            status_cell = ws.cell(row=row, column=8, value=status)
            if status == "OUT OF STOCK":
                status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                status_cell.font = Font(color='FFFFFF', bold=True)
            elif status == "LOW":
                status_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            
            ws.cell(row=row, column=9, value="Yes" if item.get('has_pending_po') else "No")
        
        self._auto_adjust_columns(ws)
        wb.save(filepath)
        return filepath
    
    def generate_dashboard_export(
        self,
        dashboard_data: Dict[str, Any],
        title: str = "Dashboard Export"
    ) -> str:
        """Generate complete dashboard data export."""
        filename = self.generate_filename("dashboard_export", "xlsx")
        filepath = self.get_report_path(filename)
        
        wb = Workbook()
        
        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        row = 4
        ws.cell(row=row, column=1, value="PO Summary").font = Font(bold=True)
        row += 1
        
        po_summary = dashboard_data.get('po_summary', {})
        for key, value in po_summary.items():
            if not isinstance(value, (list, dict)):
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=str(value))
                row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Material Summary").font = Font(bold=True)
        row += 1
        
        material_summary = dashboard_data.get('material_summary', {})
        for key, value in material_summary.items():
            if not isinstance(value, (list, dict)):
                ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=str(value))
                row += 1
        
        self._auto_adjust_columns(ws)
        wb.save(filepath)
        return filepath


# Singleton instances
pdf_generator = PDFReportGenerator()
excel_generator = ExcelReportGenerator()
